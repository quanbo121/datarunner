# -*-coding:utf-8-*
import openpyxl
import os
''' 此处的文件名配置后续会放到配置文件里'''
filename = "case1.xlsx"
newfilename = os.path.dirname(os.path.dirname(__file__))+"/Case/"+filename

class HandExcel:
    def load_excel(self):
        '''加载文件'''
        open__excel = openpyxl.load_workbook(newfilename)
        return open__excel

    def get_sheet_data(self,index=None):
        '''加载sheet页'''
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0

        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self,row,cols):
        '''获取某一个单元格内容'''
        data = self.get_sheet_data().cell(row=row,column=cols).value
        return data

    def get_rows(self,index=None):
        '''获取行数'''
        row = self.get_sheet_data(index).max_row
        return row

    def get_rows_value(self,row):
        '''获取某一用例的内容'''
        row_list = []
        '''此处遍历对象，然后把对象的值加入到列表中'''
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    '''给excle添加内容的方法'''
    def excel_write_data(self,row,cols,value):
        '''加载excel文件'''
        wb = self.load_excel()
        '''激活文件'''
        wr = wb.active
        '''写入文件'''
        wr.cell(row,cols,value)
        '''保存文件'''
        wb.save(newfilename)

    '''获取某一整列数据'''
    def get_colums_value(self,key = None):
        colums_list = []
        if key ==None:
            key = "A"
        colums_listt_data  = self.get_sheet_data()[key]
        for i in colums_listt_data:
            colums_list.append(i.value)
        return colums_list
    '''获取行号'''
    def get_rows_number(self,case_id):
        num = 0
        cols_data = self.get_colums_value()
        for col_data in cols_data:
            num = num+1
            if case_id == col_data:
                return num

    '''获取excel里面所有用例的数据'''
    def get_excel_data(self):
        data_list = []
        '''循环总行数的次数'''
        for i in range(self.get_rows()):
            '''当i为0 则要给列表里添加i+2行的数据'''
            data_list.append(self.get_rows_value(i+2))
            '''返回列表中有用例字段的部分'''
        return data_list[0:self.get_rows()-1]
'''实例化该类，方便后续调用'''
excel_data = HandExcel()
if __name__ == '__main__':
    print(excel_data.get_excel_data())
